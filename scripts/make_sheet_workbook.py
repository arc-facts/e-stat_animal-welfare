#!/usr/bin/env python3
"""手で管理するCSVを1冊のExcelブックにまとめる（Googleスプレッドシート用の雛形）。

    pip install openpyxl
    python scripts/make_sheet_workbook.py

生成された .xlsx を Googleドライブにアップロードし、
「Googleスプレッドシートとして開く」で変換すると、タブ構成そのままのシートになる。
以後の運用はシート側で行い、scripts/fetch_sheets.py が取り込む。

e-Stat API から自動取得している系列（livestock / slaughter / population / fte）は
シートに載せない。自動更新と手編集がぶつかるため。
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUT = ROOT / "build" / "animal-welfare-data.xlsx"

TABS = [
    ("cagefree", "日本のケージフリー割合の推計（出典ごと）"),
    ("cagefree_types", "ケージフリーの飼養形態別内訳"),
    ("cagefree_world", "各国のケージフリー割合"),
    ("cagefree_sea", "東南アジア6か国の規模と企業の宣言"),
    ("space_per_hen", "1羽当たり飼養面積（実面積比の図）"),
    ("pain_hours", "ケージフリー移行で減る痛みの時間"),
    ("broiler_density", "ブロイラーの飼養密度"),
    ("sow_cycle", "母豚の繁殖サイクル"),
    ("euthanasia", "犬猫の殺処分数"),
    ("hen_timeline", "ケージ飼育された採卵鶏の一生（年表）"),
    ("broiler_timeline", "ブロイラーの一生（年表）"),
    ("broiler_growth", "ブロイラーの品種別の成長（Zuidhof 2014）"),
    ("claims", "主張の台帳（ファクトチェック用・同期しない）"),
]

GUIDE = [
    ("日本の畜産動物はいま — データ管理シート", "head"),
    ("", None),
    ("このシートを編集すると、GitHub Actions が取り込んで公開サイトに反映します。", None),
    ("https://arc-facts.github.io/e-stat_animal-welfare/", None),
    ("", None),
    ("使い方", "h2"),
    ("1. 下のタブの数値を直接書き換える", None),
    ("2. 保存は不要（Googleスプレッドシートは自動保存）", None),
    ("3. 反映は6時間ごとに自動。すぐ反映したいときは GitHub の Actions で", None),
    ("   「Update data from spreadsheet」を Run workflow から手動実行", None),
    ("", None),
    ("守っていただきたいこと", "h2"),
    ("・1行目の見出し（英語の列名）は変えない。並び順も変えない", None),
    ("　→ 変わっていると取り込みが中止され、サイトは前の状態のまま保たれます", None),
    ("・タブ名は変えない", None),
    ("・数値のセルに単位や記号を入れない（「1,234羽」ではなく 1234）", None),
    ("・空欄は「データなし」として扱われます。0 とは意味が違います", None),
    ("・行の追加・削除は自由。ただし行数が半分以下に減ると安全のため取り込みを止めます", None),
    ("・メモ用のタブを自由に足して構いません（設定にないタブは無視されます）", None),
    ("", None),
    ("ファクトチェックの進め方", "h2"),
    ("claims タブが、サイト上の検証可能な主張をひとつずつ並べた台帳です。", None),
    ("1つの資料が複数の主張の根拠になっている場合も、主張ごとに1行に分けてあります。", None),
    ("・checked … 確認できたらチェックを入れる", None),
    ("・checked_by / checked_on … 誰がいつ確認したか", None),
    ("・memo … 食い違いや気づいた点。数値が違っていたらここに書いて共有する", None),
    ("データの各タブにも claim_id があり、その数値がどの主張に当たるかを辿れます。", None),
    ("source_url が空欄の行は、まだ原典を特定できていません。優先して探してください。", None),
    ("※ claims タブはリポジトリに同期しません（氏名が公開されないようにするため）。", None),
    ("", None),
    ("出典の2列について", "h2"),
    ("各タブの右端に source_text と source_url があります。", None),
    ("・source_text … その数値の根拠になった箇所を、原典の文言のまま引用する", None),
    ("・source_url  … 人手で確かめるときに、その原典に直接たどり着けるURL", None),
    ("数値を書き換えたら、この2列も必ず一緒に更新してください。", None),
    ("PDFなどダウンロードが要る資料は、共有フォルダ「AWダッシュボード」に置いて", None),
    ("そのファイルのURLを貼ってください。", None),
    ("", None),
    ("このシートで扱わないデータ", "h2"),
    ("飼養数・屠殺数・人口・労働時間（livestock / slaughter / population / fte）は", None),
    ("e-Stat API から自動取得しているため、ここにはありません。", None),
    ("手で直す必要がある場合は、リポジトリの data/ 内のCSVを直接編集してください。", None),
    ("", None),
    ("タブ一覧", "h2"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "はじめに"
    fonts = {"head": Font(bold=True, size=13), "h2": Font(bold=True)}

    r = 1
    for text, kind in GUIDE:
        cell = ws.cell(row=r, column=1, value=text)
        if kind:
            cell.font = fonts[kind]
        r += 1
    for name, desc in TABS:
        ws.cell(row=r, column=1, value=name).font = Font(name="Courier New")
        ws.cell(row=r, column=2, value=desc)
        r += 1
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 46
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top")

    header_fill = PatternFill("solid", fgColor="E6E5D9")
    for name, _desc in TABS:
        rows = list(csv.reader(open(DATA_DIR / f"{name}.csv", encoding="utf-8")))
        s = wb.create_sheet(name)
        for i, row in enumerate(rows, start=1):
            for j, v in enumerate(row, start=1):
                if i > 1 and v not in ("", None):
                    try:
                        v = int(v) if v.lstrip("-").isdigit() else float(v)
                    except ValueError:
                        pass
                cell = s.cell(row=i, column=j, value=v)
                if i == 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        s.freeze_panes = "A2"
        for j in range(1, len(rows[0]) + 1):
            width = max(len(str(row[j - 1])) if j - 1 < len(row) else 0 for row in rows)
            s.column_dimensions[get_column_letter(j)].width = min(46, max(12, width + 3))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"{OUT} を生成しました（{len(wb.sheetnames)} タブ）。")


if __name__ == "__main__":
    main()
